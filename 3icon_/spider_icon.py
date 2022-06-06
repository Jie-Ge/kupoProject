# !/usr/bin/python
# -*- coding: utf-8 -*-
'''
爬货币网站：https://tronscan.org/#/tokens/list
'''
import requests
from fake_useragent import UserAgent
import random
import threading
import xlsxwriter as xw
import os
import openpyxl as xl
from retrying import retry
import time

ua = UserAgent()

url = 'https://apilist.tronscan.org/api/tokens/overview'
start = -20
threadLock = threading.Lock()
fail_url = []
fail_img_url = []
not_img_url_count = 0
duplicated_img_count = 0
save_img_error = []

## 出错时保存 start ，下次启动接着执行
class IconThread(threading.Thread):
    def __init__(self, worksheet1, workbook):
        threading.Thread.__init__(self)
        self.worksheet1 = worksheet1
        self.workbook = workbook

    # @retry(stop_max_attempt_number=3)
    def run(self) -> None:
        global start
        global fail_url
        # total = 74060
        while start < 500:
            start += 20
            self.request_page()

            time.sleep(2+random.random())

    def request_page(self):
        print('当前存活线程数：', threading.active_count())

        params = {
            'start': start,
            'limit': 20,
            'verifier': 'all',
            'order': 'desc',
            'filter': 'trc20',
            'sort': 'marketcap',
            'order_current': 'descend'
        }
        page_num = start / 20 + 1
        print(f'正在获取第 {page_num} 页数据...{start}')

        count = 3  # 重试次数
        while count > 0:
            headers = {
                'Accept': 'application/json, text/plain, */*',
                'Connection': 'keep-alive',
                'Host': 'apilist.tronscan.org',
                'Origin': 'https://tronscan.org',
                'Referer': 'https://tronscan.org/',
                'User-Agent': random.choice(ua.data_browsers['chrome'])
            }
            try:
                res = requests.get(url, headers=headers, params=params, timeout=5)  # 这里可能会报连接超时等错误
            except Exception as e:
                print('请求失败！！！')
                if count == 1:
                    fail_url.append(url)  # 请求3次，均失败，则加入失败队列
                    # 同：traceback.print_exc(e)
                    print(e)

            else:
                if '/busy' in res.url:
                    print('请求过快，等待5秒...')
                    time.sleep(5)
                    count -= 1
                    continue

                if res.status_code == 200:
                    self.parse_data(res)
                    break
            count -= 1

    def parse_data(self, response):
        data = response.json()
        for tokens in data['tokens']:
            info_list = []
            info_list.append(tokens['name'])
            info_list.append(tokens['abbr'])
            info_list.append(tokens['nrOfTokenHolders'])
            info_list.append(tokens['contractAddress'])
            info_list.append(tokens['decimal'])
            info_list.append(tokens['projectSite'])
            info_list.append(tokens['whitePaper'])
            imgUrl = tokens['imgUrl']

            self.save_detail(info_list)
            self.down_img(imgUrl)

    def down_img(self, imgUrl):
        global not_img_url_count
        if imgUrl == '' or not imgUrl.startswith('http'):
            # print('img url 非法 ！！！')
            not_img_url_count += 1
            return None

        count = 3  # 重试次数
        while count > 0:
            headers = {
                'User-Agent': random.choice(ua.data_browsers['chrome'])
            }
            try:
                res = requests.get(imgUrl, headers=headers, timeout=5)  # 这里可能会报连接超时等错误
            except Exception as e:
                print('imgUrl请求失败！！！')
                if count == 1:
                    fail_img_url.append(imgUrl)  # 请求3次，均失败，则加入失败队列
                    # 同：traceback.print_exc(e)
                    print(e)

            else:
                if '/busy' in res.url:
                    print('请求过快，等待5秒...')
                    time.sleep(5)
                    count -= 1
                    continue

                if res.status_code == 200:
                    self.save_img(res)
                    break
                else:
                    fail_img_url.append(imgUrl)
                    break
            count -= 1

    def save_img(self, response):
        global duplicated_img_count
        global save_img_error

        img_data = response.content
        print(response.url)
        title = response.url.split('/')[-1]
        title = title.split('?')[0]

        # 统计重复图片的数量
        if os.path.exists(f'./icon_data/icon/{title}'):
            duplicated_img_count += 1
            return None

        threadLock.acquire()
        try:
            with open(f'./icon_data/icon/{title}', 'wb') as f:
                f.write(img_data)
        except Exception as e:
            save_img_error.append(response.url)
            print(e)
        threadLock.release()

    def save_detail(self, data):
        global index

        threadLock.acquire()
        try:
            self.worksheet1.append(data)
            # 这里不能保存文件，否则多线程会很慢
            # self.workbook.save('./icon_data/token_detail1.xlsx')
        except Exception as e:
            print(e)
        threadLock.release()


def main():
    if not os.path.exists('icon_data'):
        os.mkdir('icon_data')

    if not os.path.exists('./icon_data/icon'):
        os.mkdir('./icon_data/icon')

    # workbook = ''
    # worksheet1 = ''
    if os.path.exists('./icon_data/token_detail1.xlsx'):
        workbook = xl.load_workbook('./icon_data/token_detail1.xlsx')
        worksheet1 = workbook["sheet1"]
    else:
        workbook = xl.Workbook()
        worksheet1 = workbook.create_sheet("sheet1", 0)  # 在第0个位置创建
        head = ['name', 'abbr', 'Token_Holders', 'Contract', 'Decimal_Places', 'Official_Website', 'White_Pape']
        worksheet1.append(head)

    threads = []
    for i in range(16):
        t = IconThread(worksheet1, workbook)
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    workbook.save('./icon_data/token_detail1.xlsx')

    print(f'请求失败的url个数：{len(fail_url)}')
    print(f'请求失败的img url个数：{len(fail_img_url)}\n', fail_img_url)
    print(f'不是img url的个数：{not_img_url_count}')
    print(f'重复图片数量：{duplicated_img_count}')
    print(f'保存失败的img url个数：{len(save_img_error)}\n', save_img_error)


if __name__ == '__main__':
    start_time = time.time()
    main()
    end_time = time.time()
    print('程序结束！！！，用时: ', end_time-start_time)

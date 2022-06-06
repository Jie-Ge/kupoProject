# !/usr/bin/python
# -*- coding: utf-8 -*-
'''
爬取：https://bscscan.com/tokens?ps=100&p=1
'''

import threading
import time
import requests
from fake_useragent import UserAgent
import random
from lxml import etree
import os
import openpyxl as xl

ua = UserAgent()

page_num = 0
url_num = 0  # url计数，用于显示进度
fail_url = []
threadLock = threading.Lock()

class Icon2Thread(threading.Thread):
    def __init__(self, workbook, worksheet1):
        threading.Thread.__init__(self)
        self.workbook = workbook
        self.worksheet1 = worksheet1

    def run(self) -> None:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'user-agent': random.choice(ua.data_browsers['chrome'])
        }
        url = 'https://bscscan.com/tokens'
        global page_num
        while page_num < 4:
            page_num += 1
            params = {
                'ps': 100,
                'p': page_num
            }
            print(f'正在请求第 {page_num} 个列表页...')
            try:
                res = requests.get(url, headers=headers, params=params, timeout=5)
            except Exception as e:
                print('列表页请求失败', res.url)
                print(e)
            else:
                if res.status_code == 200:
                    self.parse_page_url(res)

    def parse_page_url(self, response):
        page = etree.HTML(response.text)
        h3_tags = page.xpath('//h3')

        if len(h3_tags):
            for h3_tag in h3_tags:
                href = h3_tag.xpath('./a/@href')
                if len(href):
                    page_href = href[0]
                    if page_href.startswith('/token'):
                        self.request_page(page_href, response.url)

    def request_page(self, url, referer):
        global url_num
        url_num += 1
        print(f'正在请求第 {url_num}/382 个url：', threading.active_count())
        count = 3  # 重试次数
        url = 'https://bscscan.com' + url
        while count > 0:
            headers = {
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                'referer': referer,
                'user-agent': random.choice(ua.data_browsers['chrome'])
            }
            try:
                res = requests.get(url, headers=headers, timeout=5)  # 这里可能会报连接超时等错误
            except Exception as e:
                print('请求失败！！！')
                if count == 1:
                    fail_url.append(url)  # 请求3次，均失败，则加入失败队列
                    # 同：traceback.print_exc(e)
                    print(e)

            else:
                if '/busy' in res.url:
                    print('请求过快，等待5秒...')
                    time.sleep(5)
                    count -= 1
                    continue

                if res.status_code == 200:
                    self.parse_data(res)
                    break
            count -= 1

    def parse_data(self, response):
        title = ''
        img_url = ''
        Holders = ''
        Contract = ''
        Decimals = ''
        Official_Site = ''

        page = etree.HTML(response.text)

        img_label = page.xpath('//h1/img/@src')
        if len(img_label):
            img_url = 'https://bscscan.com' + img_label[0]

        title_label = page.xpath('//h1//span[@class="text-secondary small"]/text()')
        if len(title_label):
            title = title_label[0]

        Holders_label = page.xpath('//div[@id="ContentPlaceHolder1_tr_tokenHolders"]//div[@class="mr-3"]/text()')
        if len(Holders_label):
            Holders = Holders_label[0].split(' ')[0]

        Contract_label = page.xpath('//div[@class="d-flex clipboard-hover"]/a[1]/text()')
        if len(Contract_label):
            Contract = Contract_label[0]

        Decimals_label = page.xpath('//div[@id="ContentPlaceHolder1_trDecimals"]//div[@class="col-md-8"]/text()')
        if len(Decimals_label):
            Decimals = Decimals_label[0]

        Official_Site_label = page.xpath('//div[@id="ContentPlaceHolder1_tr_officialsite_1"]//div[@class="col-md-8"]/a/text()')
        if len(Official_Site_label):
            Official_Site = Official_Site_label[0]

        detail_info = []
        detail_info.append(title)
        detail_info.append(Holders)
        detail_info.append(Contract)
        detail_info.append(Decimals)
        detail_info.append(Official_Site)

        if detail_info[2] == '':
            print(detail_info, response.url)

        self.save_detail(detail_info)
        self.down_img(img_url)

    def save_detail(self, data):
        print(data)


        threadLock.acquire()
        try:
            self.worksheet1.append(data)
        except Exception as e:
            print(e)
        threadLock.release()

    def down_img(self, url):
        if url == '' or not url.startswith('http'):
            return None
        headers = {
            'User-Agent': random.choice(ua.data_browsers['chrome'])
        }
        try:
            res = requests.get(url, headers=headers, timeout=5)
        except:
            print('img 下载失败!!!')
            return None
        else:
            if res.status_code == 200:
                self.save_img(res)

    def save_img(self, res):
        print(res.url)
        title = res.url.split('/')[-1]
        title = title.split('?')[0]

        threadLock.acquire()
        try:
            with open(f'./icon2_data/icon2/{title}', 'wb') as f:
                f.write(res.content)
        except:
            print('img 保存失败！！！')
        threadLock.release()


def main():
    if not os.path.exists('./icon2_data'):
        os.mkdir('./icon2_data')

    if not os.path.exists('./icon2_data/icon2'):
        os.mkdir('./icon2_data/icon2')

    if os.path.exists('icon2_data/token_detail1.xlsx'):
        workbook = xl.load_workbook('icon2_data/token_detail1.xlsx')
        worksheet1 = workbook['sheet1']
    else:
        workbook = xl.Workbook()
        worksheet1 = workbook.create_sheet('sheet1', 0)  # 在第0个位置创建
        head = ['title', 'Holders', 'Contract', 'Decimals', 'Official_Site']
        worksheet1.append(head)


    threads = []
    for i in range(10):
        t = Icon2Thread(workbook, worksheet1)
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    workbook.save('./icon2_data/token_detail1.xlsx')
    print(f'{len(fail_url)}个url请求失败：', fail_url)


if __name__ == '__main__':
    main()
